import io
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from toolbox_py import get_logger

from app.document_schema import DOCUMENT_SCHEMA
from app.models.document import ExtractedDocument
from app.utils.exceptions import ExcelGenerationError

logger = get_logger(__name__)

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def _build_excel_columns() -> list[tuple[str, list[str]]]:
    """Derive ordered Excel column definitions from :data:`~app.document_schema.DOCUMENT_SCHEMA`.

    Iterates the schema in definition order.  Fields with the same label across
    different document types are merged into one column: when building a data
    row the first non-``None`` value among the merged field keys wins.

    Internal-only fields (``FieldSpec.label is None``) are excluded from the
    output.

    Returns:
        Ordered list of ``(label, [field_keys])`` pairs.
    """
    column_keys: OrderedDict[str, list[str]] = OrderedDict()
    for doc_spec in DOCUMENT_SCHEMA.values():
        for field_spec in doc_spec.fields:
            if field_spec.label is None:
                continue
            if field_spec.label not in column_keys:
                column_keys[field_spec.label] = []
            if field_spec.key not in column_keys[field_spec.label]:
                column_keys[field_spec.label].append(field_spec.key)
    return list(column_keys.items())


# Derived at import time from DOCUMENT_SCHEMA — never hardcoded.
_EXCEL_COLUMNS: list[tuple[str, list[str]]] = _build_excel_columns()
_HEADERS: list[str] = (
    ["Arquivo", "Tipo de Documento"] + [label for label, _ in _EXCEL_COLUMNS] + ["Texto Extraído"]
)


class ExcelService:
    """Service for generating Excel reports from extracted document data."""

    def generate(self, documents: list[ExtractedDocument]) -> bytes:
        """Create an Excel workbook from a list of extracted documents.

        The set of columns and their order are derived entirely from
        :data:`~app.document_schema.DOCUMENT_SCHEMA`.

        Args:
            documents: List of :class:`~app.models.document.ExtractedDocument`
                objects to include in the report.

        Returns:
            Raw bytes of the generated ``.xlsx`` file.

        Raises:
            ExcelGenerationError: If Excel generation fails.
        """
        try:
            logger.debug(
                "Generating Excel report",
                extra={"document_count": len(documents)},
            )
            wb = Workbook()
            ws = wb.active
            if ws is None:  # pragma: no cover – new Workbook always has an active sheet
                raise ExcelGenerationError("Workbook has no active worksheet")
            ws.title = "Documentos"

            # Write styled header row
            for col_idx, header in enumerate(_HEADERS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = _HEADER_ALIGNMENT

            # Write data rows
            for row_idx, doc in enumerate(documents, start=2):
                fields = doc.extracted_fields
                row_data: list[str | None] = [doc.filename, doc.document_type.value]

                for _label, keys in _EXCEL_COLUMNS:
                    value = next(
                        (fields[k] for k in keys if k in fields and fields[k] is not None),
                        None,
                    )
                    row_data.append(value)

                row_data.append(doc.raw_text)

                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-fit column widths based on header length
            for col_idx, header in enumerate(_HEADERS, start=1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(len(header) + 4, 16)

            buffer = io.BytesIO()
            wb.save(buffer)
            file_bytes = buffer.getvalue()
            logger.debug(
                "Excel report generated",
                extra={"document_count": len(documents), "file_size_bytes": len(file_bytes)},
            )
            return file_bytes

        except Exception as exc:
            logger.error("Excel generation failed: %s", exc)
            raise ExcelGenerationError(f"Failed to generate Excel file: {exc}") from exc
