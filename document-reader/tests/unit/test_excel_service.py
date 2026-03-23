import openpyxl
import pytest

from app.models.document import DocumentType, ExtractedDocument
from app.services.excel_service import ExcelService


@pytest.fixture
def svc() -> ExcelService:
    return ExcelService()


def _make_doc(
    filename: str = "test.jpg", doc_type: DocumentType = DocumentType.RG
) -> ExtractedDocument:
    return ExtractedDocument(
        filename=filename,
        document_type=doc_type,
        raw_text="Some raw text",
        extracted_fields={"name": "João Silva", "rg_number": "12.345.678-9"},
    )


def test_generate_returns_bytes(svc: ExcelService) -> None:
    result = svc.generate([_make_doc()])
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_generate_creates_valid_excel(svc: ExcelService) -> None:
    result = svc.generate([_make_doc()])
    import io

    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert "Documentos" in wb.sheetnames


def test_generate_correct_headers(svc: ExcelService) -> None:
    result = svc.generate([_make_doc()])
    import io

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Documentos"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, 14)]
    assert "Arquivo" in headers
    assert "Tipo de Documento" in headers
    assert "Número RG" in headers
    assert "Texto Extraído" in headers


def test_generate_with_multiple_documents(svc: ExcelService) -> None:
    docs = [
        _make_doc("doc1.jpg", DocumentType.RG),
        _make_doc("doc2.jpg", DocumentType.PIS),
        _make_doc("doc3.jpg", DocumentType.PROOF_OF_ADDRESS),
    ]
    result = svc.generate(docs)
    import io

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Documentos"]
    # header row + 3 data rows = 4 total rows
    assert ws.max_row == 4
